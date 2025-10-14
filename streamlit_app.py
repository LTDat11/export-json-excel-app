# -*- coding: utf-8 -*-
import json
import io
import os
import csv
import threading
from datetime import datetime
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- Config ---
DATA_DIR = "shared_notes"
LOG_FILE = "export_logs.csv"
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs("logs", exist_ok=True)
log_lock = threading.Lock()

# --- Safe log writing ---
def write_log(user, filename, total_shirt, total_films):
    log_path = os.path.join("logs", LOG_FILE)
    with log_lock:
        new_file = not os.path.exists(log_path)
        with open(log_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if new_file:
                writer.writerow(["user", "file_name", "total_shirt", "total_films", "timestamp"])
            writer.writerow([user, filename, total_shirt, total_films, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

# --- Excel export function ---
def export_to_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws["A1"], ws["B1"], ws["C1"] = "FILE", "DATE", "TYPE"
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws.merge_cells("C1:C2")
    for c in ["A1", "B1", "C1"]:
        ws[c].alignment = Alignment(horizontal="center", vertical="center")
        ws[c].font = Font(bold=True)
        ws[c].border = border

    headers = ["ORDER ID", "ITEM", "F/B", "SHIRT TYPE", "QUANT.", "COLOR", "SIZE", "Approved", "Note"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    current_row = 5
    orders = {}
    for item in data:
        order_id = item.get("order_external_id", "")
        idx_item = item.get("index_item", "")
        orders.setdefault(order_id, {}).setdefault(idx_item, []).append(item)

    total_all_films = 0
    total_all_shirts = 0

    for order_id, groups in orders.items():
        for idx_item, group_items in sorted(groups.items(), key=lambda x: int(x[0]) if str(x[0]).isdigit() else x[0]):
            item_count = len(group_items)
            labels = sorted({x.get("label", "").strip() for x in group_items if x.get("label", "")})
            fb_value = "/".join(labels) if labels else ""
            first = group_items[0]
            shirt_type = first.get("product_name", "").upper()
            color = first.get("product_color", "").strip()
            size = first.get("product_size", "").strip()

            row_vals = [order_id, item_count, fb_value, shirt_type, "1", color, size]
            for col, val in enumerate(row_vals, 1):
                cell = ws.cell(current_row, col, val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.cell(current_row, 8, "").border = border
            ws.cell(current_row, 9, "").border = border

            total_all_films += item_count
            total_all_shirts += 1
            current_row += 1

    ws.cell(current_row, 1, "...")
    for c in range(1, 10):
        ws.cell(current_row, c).border = border
        ws.cell(current_row, c).alignment = Alignment(horizontal="center", vertical="center")
    current_row += 1

    ws.cell(current_row, 1, "TOTAL FILMS")
    ws.cell(current_row, 2, total_all_films)
    ws.cell(current_row, 4, "TOTAL SHIRT")
    ws.cell(current_row, 5, total_all_shirts)
    for c in range(1, 10):
        ws.cell(current_row, c).border = border
        ws.cell(current_row, c).alignment = Alignment(horizontal="center", vertical="center")

    widths = [18, 8, 18, 20, 10, 16, 12, 10, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, total_all_shirts, total_all_films


# --- UI ---
st.set_page_config(page_title="JSON ➜ Excel Tool (Multi-user + Admin)", layout="wide")
st.title("🧾 JSON ➜ Excel Export Tool (4 người nhập + 1 quản lý)")

users = ["Tiên", "Hải", "Dung", "Sơn"]
tabs = st.tabs(["👤 Tiên", "👤 Hải", "👤 Dung", "👤 Sơn", "🧑‍💼 Admin"])

# --- User tabs ---
for i, user in enumerate(users):
    with tabs[i]:
        st.subheader(f"📋 Notepad của {user}")
        json_path = os.path.join(DATA_DIR, f"{user}.json")

        existing = ""
        if os.path.exists(json_path):
            with open(json_path, "r", encoding="utf-8") as f:
                existing = f.read()

        json_input = st.text_area(f"Dán JSON ({user}):", value=existing, height=250, key=f"input_{user}")

        if st.button(f"💾 Lưu JSON ({user})", key=f"save_{user}"):
            text = json_input.strip()
            if not text:
                # Nếu trống → ghi [] để đánh dấu "đã xóa"
                with open(json_path, "w", encoding="utf-8") as f:
                    f.write("[]")
                st.warning(f"⚠️ {user} đã xóa toàn bộ dữ liệu (file sẽ được ghi [] để tránh lỗi).")
            else:
                try:
                    data = json.loads(text)
                    if not isinstance(data, list):
                        st.error("❌ JSON phải là list (danh sách) các object.")
                    else:
                        with open(json_path, "w", encoding="utf-8") as f:
                            json.dump(data, f, ensure_ascii=False, indent=2)
                        st.success(f"✅ Đã lưu JSON cho {user} (file cũ bị ghi đè).")
                except json.JSONDecodeError as e:
                    st.error(f"❌ JSON không hợp lệ!\n{e}")


# --- Admin tab ---
with tabs[-1]:
    st.subheader("🧑‍💼 Admin - Quản lý & Export")

    # kiểm tra trạng thái từng người
    status_list = []
    available = []
    for u in users:
        path = os.path.join(DATA_DIR, f"{u}.json")
        if not os.path.exists(path):
            status_list.append(f"🔴 {u}: chưa có file JSON")
        else:
            try:
                data = json.load(open(path, "r", encoding="utf-8"))
                if isinstance(data, list) and len(data) > 0:
                    status_list.append(f"🟢 {u}: có {len(data)} bản ghi")
                    available.append(u)
                else:
                    status_list.append(f"🟡 {u}: file rỗng hoặc JSON không hợp lệ")
            except Exception:
                status_list.append(f"🔴 {u}: lỗi khi đọc JSON")

    st.markdown("### 📊 Trạng thái người dùng:")
    for s in status_list:
        st.write(s)

    selected_users = st.multiselect("Chọn người cần export:", available, default=available)

    if st.button("📤 Export Excel"):
        combined = []
        for u in selected_users:
            path = os.path.join(DATA_DIR, f"{u}.json")
            with open(path, "r", encoding="utf-8") as f:
                user_data = json.load(f)
                if isinstance(user_data, list):
                    combined.extend(user_data)

        if not combined:
            st.warning("⚠️ Không có dữ liệu hợp lệ để export.")
        else:
            buffer, total_shirt, total_films = export_to_excel(combined)
            filename = "_".join(selected_users)
            final_name = f"{filename}_TOTAL_SHIRT_{total_shirt}_TOTAL_FILMS_{total_films}.xlsx"
            write_log("Admin", final_name, total_shirt, total_films)

            st.success(f"✅ Xuất Excel thành công từ {', '.join(selected_users)}")
            st.download_button(
                "⬇️ Tải Excel",
                data=buffer,
                file_name=final_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    log_path = os.path.join("logs", LOG_FILE)
    if os.path.exists(log_path):
        st.divider()
        st.subheader("📜 Lịch sử xuất file")
        df = pd.read_csv(log_path)
        st.dataframe(df, use_container_width=True)
